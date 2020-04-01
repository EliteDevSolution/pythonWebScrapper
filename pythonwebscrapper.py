from selenium import webdriver
import time
import openpyxl
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
import requests
from lxml import html

class Scraper:
    def __init__(self):
        '''
        When we initialize a scraper object, we need to create our webdriver
        instance
        '''
        # Please note that this is the directory of my webdriver instance
        self.browser = webdriver.Chrome()
        self.browser.get('https://www.ibba.org')

    def search(self, search_area):
        '''
        This method is used to fetch the ibba.org address and find the search
        bar element to enter in a search that we are concerned with, in this
        case, we are searching for all brokers in the Los Angeles area
        '''
        search = self.browser.find_element_by_id('locator-address2')
        search.clear()
        search.send_keys(search_area)
        #search.send_keys("Los Angeles")
        search.send_keys(u'\ue007')
        time.sleep(1)

    def get_num_of_brokers(self):
        '''
        Fetches the number of brokers displayed at the top of a screen after a
        search
        '''
        try:
            num_of_brokers = self.browser.find_element_by_xpath('//*[@id="CPRLocator"]/div[1]').text.split()[2]
        except :
            num_of_brokers = '0'

        return int(num_of_brokers)

    def more_details(self, brokers):
        '''
        get url of the brokers
        '''
        details_div = self.browser.find_element_by_xpath('//*[@id="CPRLocator"]/div[2]')
        URLs = []
        for broker_index in range(1, scraper.get_num_of_brokers() + 1):
            detail = details_div.find_element_by_xpath('//*[@id="CPRLocator"]/div[2]/div[' + str(broker_index) + ']/section/div[2]')
            URL_text = detail.find_element_by_tag_name('a').get_attribute('href')
            if ('/' in URL_text) :
                URLs.append(URL_text)
        return URLs

    def get_info(self, url):
        '''
        get general information of the brokers
        '''
        #self.browser.get(url)
        #time.sleep(0.1)
        response = requests.get(url)
        time.sleep(0.5)

        #print(response.text)
        soup = BeautifulSoup(response.text, 'html.parser')

        info =  []

        #<div class="member-image">
        # name
        try:
            info += [soup.find("h1", {"class": "page-title"}).text]
        except :
            info += [""]

        #<table class="table" id="broker-data">
        # company
        try:
            info += [soup.find("tr", {"class": "company"}).text[8:].replace("\n", "")]
        except :
            info += [""]
        # address
        try:
            info += [soup.find("tr", {"class": "address"}).text[8:].replace("\n", "")]
        except :
            info += [""]
        # phone
        try:
            info += [soup.find("div", {"class": "phone"}).text.replace("\n", "")]
        except :
            info += [""]

        #<div class='gform_body'>
        # email
        try:
            info += [soup.find("input", {"id": "input_3_10"})['value'].replace("\n", "")]
            #mail_text = self.browser.find_element_by_xpath('//*[@id="input_3_10"]').get_attribute('value')
#            info += [self.browser.find_element_by_class_name("email").
#                     text[5:]]
#            info += [mail_text]
        except :
            info += [""]

        return info


    def quit(self):
        '''
        Call this method to close the browser instance
        '''
        self.browser.quit()

filepath = "brokers.xlsx"
workbook = openpyxl.Workbook()
workbook.save(filepath)

workbook = openpyxl.load_workbook(filepath)
sheet = workbook['Sheet']

# Initialize our scraper and browser instance
scraper = Scraper()

# Create the initial search, defined by our search method
search_areas = ["Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut",
                "Delaware", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", 
                "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", 
                "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New", "North",
                "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode", "Tennessee", 
                "Texas", "Utah", "Vermont", "Virginia", "Washington", "West", "Wisconsin", "Wyoming"]

row_index = 0
for search_area in search_areas:
    scraper.search(search_area)

    # get general information of the brokers
    brokers = scraper.get_num_of_brokers()
    if (brokers == 0) :
        continue

    print("please wait for getting broker urls ..." + " on " + search_area)
    detail_URLs = scraper.more_details(brokers)

    # get each broker information
    index = 0
    for detail_URL in detail_URLs:
        row_index = row_index + 1
        index = index + 1
        raw_data = scraper.get_info(detail_URL)
        print(str(index) + " of " + str(len(detail_URLs)) + " on " + search_area + " : ", raw_data)

        # write broker information in excel
        for i in range(0, len(raw_data)):
            sheet.cell(row_index, i + 1).value = raw_data[i]    
        sheet.cell(row_index, len(raw_data) + 1).value = search_area    

    workbook.save(filepath)
# Close our browser instance, we collected all the data!
scraper.quit()


